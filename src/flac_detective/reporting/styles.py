"""Styles et formatage pour rapports Excel."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Styles pour les en-têtes
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Bordures
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Couleurs pour les scores
SCORE_COLORS = {
    "good": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Jaune
    "suspect": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Orange
    "bad": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Rouge
}

# Largeurs de colonnes
COLUMN_WIDTHS = {
    "A": 60,  # Chemin complet
    "B": 40,  # Nom fichier
    "C": 15,  # Score
    "D": 50,  # Raison
    "E": 18,  # Fréquence
    "F": 12,  # Sample Rate
    "G": 10,  # Bit Depth
    "H": 25,  # Encodeur
    "I": 35,  # Problème Durée
    "J": 18,  # Durée Métadonnées
    "K": 18,  # Durée Réelle
}


def get_score_style(score: int) -> tuple[PatternFill, Font | None]:
    """Retourne le style approprié selon le score.

    Args:
        score: Score de qualité (0-100).

    Returns:
        Tuple (fill, font) pour le style de la cellule.
    """
    if score >= 70:
        return SCORE_COLORS["good"], None
    elif score >= 50:
        return SCORE_COLORS["suspect"], None
    else:
        return SCORE_COLORS["bad"], Font(color="FFFFFF", bold=True)
