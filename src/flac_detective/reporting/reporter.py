"""Générateur de rapports Excel professionnels."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .statistics import calculate_statistics, filter_suspicious
from .styles import (
    BORDER,
    COLUMN_WIDTHS,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    SCORE_COLORS,
    get_score_style,
)

logger = logging.getLogger(__name__)


class ExcelReporter:
    """Génération du rapport Excel professionnel."""

    def __init__(self):
        """Initialise le générateur de rapports."""
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore
        if self.ws is None:
            self.ws = self.wb.create_sheet()
        self.ws.title = "Fichiers Suspects"

    def generate_report(self, results: List[Dict], output_file: Path):
        """Génère le rapport Excel avec les résultats.

        Args:
            results: Liste des résultats d'analyse.
            output_file: Chemin du fichier Excel de sortie.
        """
        # Filtrer uniquement les fichiers suspects (score < 90)
        suspicious = filter_suspicious(results)

        # Générer la feuille de données
        self._write_data_sheet(suspicious)

        # Ajout d'une feuille de résumé
        self._add_summary_sheet(results, suspicious)

        # Sauvegarde
        self.wb.save(output_file)
        logger.info(f"✅ Rapport Excel généré: {output_file}")

    def _write_data_sheet(self, suspicious: List[Dict]):
        """Écrit la feuille de données des fichiers suspects.

        Args:
            suspicious: Liste des fichiers suspects.
        """
        # En-têtes
        headers = [
            "Chemin Complet",
            "Nom du Fichier",
            "Score FLAC (%)",
            "Raison du Doute",
            "Fréquence Coupure (Hz)",
            "Sample Rate",
            "Bit Depth",
            "Encodeur",
            "Problème Durée",
            "Durée Métadonnées",
            "Durée Réelle",
        ]

        # Écriture des en-têtes
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Écriture des données
        for row_idx, result in enumerate(suspicious, start=2):
            self._write_result_row(row_idx, result)

        # Ajustement automatique des largeurs de colonnes
        for col_letter, width in COLUMN_WIDTHS.items():
            self.ws.column_dimensions[col_letter].width = width

        # Figer la première ligne
        self.ws.freeze_panes = "A2"

    def _write_result_row(self, row_idx: int, result: Dict):
        """Écrit une ligne de résultat.

        Args:
            row_idx: Index de la ligne.
            result: Résultat d'analyse.
        """
        # Chemin complet
        cell = self.ws.cell(row=row_idx, column=1)
        cell.value = result["filepath"]
        cell.border = BORDER

        # Nom du fichier
        cell = self.ws.cell(row=row_idx, column=2)
        cell.value = result["filename"]
        cell.border = BORDER

        # Score avec code couleur
        cell = self.ws.cell(row=row_idx, column=3)
        cell.value = result["score"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        fill, font = get_score_style(result["score"])
        cell.fill = fill
        if font:
            cell.font = font

        # Raison
        cell = self.ws.cell(row=row_idx, column=4)
        cell.value = result["reason"]
        cell.alignment = Alignment(wrap_text=True)
        cell.border = BORDER

        # Fréquence de coupure
        cell = self.ws.cell(row=row_idx, column=5)
        cell.value = result["cutoff_freq"]
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Sample Rate
        cell = self.ws.cell(row=row_idx, column=6)
        cell.value = result["sample_rate"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Bit Depth
        cell = self.ws.cell(row=row_idx, column=7)
        cell.value = result["bit_depth"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Encodeur
        cell = self.ws.cell(row=row_idx, column=8)
        cell.value = result["encoder"]
        cell.border = BORDER

        # Problème Durée
        cell = self.ws.cell(row=row_idx, column=9)
        duration_mismatch = result.get("duration_mismatch")
        if duration_mismatch:
            cell.value = duration_mismatch
            cell.fill = SCORE_COLORS["suspect"]
            cell.font = Font(bold=True)
        else:
            cell.value = "✓ OK"
            cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Durée Métadonnées
        cell = self.ws.cell(row=row_idx, column=10)
        cell.value = result.get("duration_metadata", "N/A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

        # Durée Réelle
        cell = self.ws.cell(row=row_idx, column=11)
        cell.value = result.get("duration_real", "N/A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    def _add_summary_sheet(self, all_results: List[Dict], suspicious: List[Dict]):
        """Ajoute une feuille de résumé statistique.

        Args:
            all_results: Liste complète des résultats.
            suspicious: Liste des fichiers suspects.
        """
        ws_summary = self.wb.create_sheet("Résumé", 0)

        # Calcul des statistiques
        stats = calculate_statistics(all_results)

        # Titre
        ws_summary["A1"] = "RAPPORT D'ANALYSE FLAC"
        ws_summary["A1"].font = Font(size=16, bold=True, color="366092")

        # Date
        ws_summary["A2"] = f'Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M:%S")}'
        ws_summary["A2"].font = Font(size=10, italic=True)

        # Statistiques globales
        ws_summary["A4"] = "STATISTIQUES GLOBALES"
        ws_summary["A4"].font = Font(size=12, bold=True)

        stat_rows = [
            ("Fichiers analysés:", stats["total"], ""),
            ("Authentiques (90-100%):", stats["authentic"], stats["authentic_pct"]),
            (
                "Probablement authentiques (70-89%):",
                stats["probably_authentic"],
                stats["probably_authentic_pct"],
            ),
            ("Suspects (50-69%):", stats["suspect"], stats["suspect_pct"]),
            ("Très suspects (<50%):", stats["fake"], stats["fake_pct"]),
        ]

        row = 5
        for label, value, pct in stat_rows:
            ws_summary[f"A{row}"] = label
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = value
            if pct:
                ws_summary[f"C{row}"] = pct
                ws_summary[f"C{row}"].font = Font(italic=True)
            row += 1

        # Statistiques sur les problèmes de durée
        row += 1
        ws_summary[f"A{row}"] = "PROBLÈMES DE DURÉE (Critère Fakin' The Funk)"
        ws_summary[f"A{row}"].font = Font(size=12, bold=True)
        row += 1

        duration_stat_rows = [
            (
                "Fichiers avec décalage durée:",
                stats["duration_issues"],
                stats["duration_issues_pct"],
            ),
            (
                "Décalage critique (>1 seconde):",
                stats["duration_issues_critical"],
                stats["duration_issues_critical_pct"],
            ),
        ]

        for label, value, pct in duration_stat_rows:
            ws_summary[f"A{row}"] = label
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = value
            if pct:
                ws_summary[f"C{row}"] = pct
                ws_summary[f"C{row}"].font = Font(italic=True)
            row += 1

        # Note explicative
        row += 2
        ws_summary[f"A{row}"] = "ℹ️ Note sur les problèmes de durée :"
        ws_summary[f"A{row}"].font = Font(bold=True, color="0066CC")
        row += 1
        ws_summary[f"A{row}"] = (
            "Un décalage entre durée métadonnées et durée réelle peut indiquer :"
        )
        row += 1
        ws_summary[f"A{row}"] = "  • Fichier corrompu lors de l'encodage"
        row += 1
        ws_summary[f"A{row}"] = "  • Transcodage mal effectué (perte de samples)"
        row += 1
        ws_summary[f"A{row}"] = "  • Métadonnées erronées après édition manuelle"
        row += 1
        ws_summary[f"A{row}"] = "  • Problème lors d'un split/merge d'album"
        row += 1
        ws_summary[f"A{row}"] = "Tolérance normale : <588 samples (~13ms pour 44.1kHz)"

        # Ajustement des colonnes
        ws_summary.column_dimensions["A"].width = 45
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 15
